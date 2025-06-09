#!/usr/bin/env python3
"""
Script de depuración: imprime las primeras filas y celdas clave de cada hoja para identificar ubicaciones de cabeceras.
"""
from openpyxl import load_workbook

EXCEL_PATH = 'tax_calendar_25.xlsm'


def debug_header():
    wb = load_workbook(EXCEL_PATH, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Hoja: {sheet_name} ---")
        # Mostrar filas 1 a 10, columnas A a J
        for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10):
            vals = [str(cell.value) if cell.value is not None else '' for cell in row]
            print(vals)
    wb.close()

if __name__ == '__main__':
    debug_header()
