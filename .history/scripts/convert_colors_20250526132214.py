#!/usr/bin/env python3
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Rutas
ROOT   = os.path.dirname(os.path.dirname(__file__))
INPUT  = os.path.join(ROOT, "tax_calendar_25.xlsm")
OUTPUT = os.path.join(ROOT, "tax_calendar_25_cleaned.xlsm")

# Mapa de color ARGB → sigla
COLOR_CODE = {
    "FFFFFF": "SP",  # blanco
    "FFFF66": "SI",
    "9966FF": "RI",
    "FF5050": "SD",
    "FF99FF": "AD",
    "70AD47": "OS",
    "00B0F0": "OP",
    "996633": "HS",
    "BF8F00": "HL",
    "CC9900": "HL",
}
SIGLAS = set(COLOR_CODE.values())

def argb_to_hex(argb):
    if not argb:
        return None
    s = str(argb)
    return s[-6:].upper()

def main():
    print(f"INPUT file:  {INPUT}")
    print(f"OUTPUT file: {OUTPUT}")
    if not os.path.exists(INPUT):
        print("ERROR: no existe el archivo de entrada")
        sys.exit(1)

    try:
        wb = load_workbook(INPUT, keep_vba=True)
    except Exception as e:
        print("ERROR al abrir el workbook:", e)
        sys.exit(1)

    processed = 0
    for sheet in wb.sheetnames:
        if sheet == "SETTINGS" or sheet.startswith("CALENDAR"):
            continue
        ws = wb[sheet]
        print(f"\nProcesando hoja: {sheet}")
        processed += 1

        # Borra reglas condicionales
        try:
            ws.conditional_formatting._cf_rules = {}
        except Exception:
            pass

        cells_changed = 0
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                # 1) Si ya es sigla válida, ignorar
                if isinstance(val, str) and val.strip().upper() in SIGLAS:
                    continue

                # 2) Si tiene relleno sólido, convertir
                fill = cell.fill
                if isinstance(fill, PatternFill) and fill.fill_type == "solid":
                    hexa = argb_to_hex(fill.start_color.rgb)
                    code = COLOR_CODE.get(hexa)
                    if code:
                        cell.value = code
                        cells_changed += 1
                    # limpiar relleno en cualquier caso
                    cell.fill = PatternFill(fill_type=None)

        print(f"  celdas modificadas en '{sheet}': {cells_changed}")

    print(f"\nTotal hojas procesadas: {processed}")
    # Guardar
    try:
        wb.save(OUTPUT)
        print("Guardado correctamente en:", OUTPUT)
    except Exception as e:
        print("ERROR al guardar el archivo:", e)
        sys.exit(1)

if __name__ == "__main__":
    main()
