from openpyxl import load_workbook
from datetime import datetime, date
from typing import List, Dict, Any, Optional
import re

def argb_to_hex(argb) -> Optional[str]:
    if not argb:
        return None
    s = str(argb)
    return f"#{s[-6:]}"

EMPRESA_OVERRIDES = {
    "France": "Repsol",
    "Netherlands": "Repsol",
}

def parse_calendar(
    path_excel: str,
    target_date: Optional[date] = None
) -> List[Dict[str, Any]]:
    wb = load_workbook(path_excel, data_only=True)
    registros: List[Dict[str, Any]] = []

    for sheet in wb.sheetnames:
        if sheet == "SETTINGS" or sheet.startswith("CALENDAR"):
            continue
        ws = wb[sheet]
        empresa = EMPRESA_OVERRIDES.get(sheet, sheet)

        # 1) Encontrar fila de Mes-Año
        month_row = None
        for r in range(1, 11):
            for cell in ws[r]:
                val = cell.value
                if isinstance(val, str) and re.match(r"^[A-Za-z]+ - \d{4}$", val.strip()):
                    month_row = r
                    break
            if month_row:
                break
        if not month_row:
            continue

        # 2) Construir month_info
        month_info: Dict[int, tuple] = {}
        current_my = None
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=month_row, column=col).value
            if isinstance(val, str) and re.match(r"^[A-Za-z]+ - \d{4}$", val.strip()):
                mes_str, anyo_str = val.split(" - ")
                try:
                    anyo = int(anyo_str)
                    current_my = (mes_str, anyo)
                except ValueError:
                    current_my = None
            if current_my:
                month_info[col] = current_my

        # 3) Fila de días fija en 5
        day_row = 5
        # validación mínima
        nums = [c.value for c in ws[day_row] if isinstance(c.value, (int, float))]
        if not (len(set(nums)) >= 2 and all(1 <= int(v) <= 31 for v in nums)):
            continue

        # 4) Construir date_map
        date_map: Dict[int, date] = {}
        for col, (mes, anyo) in month_info.items():
            cell = ws.cell(row=day_row, column=col)
            if isinstance(cell.value, (int, float)):
                dia = int(cell.value)
                try:
                    fecha = datetime.strptime(f"{dia} {mes} {anyo}", "%d %B %Y").date()
                    date_map[col] = fecha
                except ValueError:
                    pass

        # 5) Si nos piden target_date, buscar la columna concreta
        if target_date:
            # invertimos date_map
            inv = {v: k for k, v in date_map.items()}
            col_for_date = inv.get(target_date)
            if not col_for_date:
                continue
            cols_to_check = [col_for_date]
        else:
            cols_to_check = list(date_map.keys())

        # 6) Recorrer filas de datos hasta antes de leyenda
        current_country = None
        current_tax = None
        for r in range(day_row + 1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            # parar si encontramos la leyenda
            if isinstance(a, str) and a.strip().lower().startswith("legend"):
                break
            b = ws.cell(row=r, column=2).value
            if isinstance(a, str) and a.strip():
                current_country = a.strip()
            if isinstance(b, str) and b.strip():
                current_tax = b.strip()
            if not current_country or not current_tax:
                continue

            # 7) Revisar solo las columnas de interés
            for col in cols_to_check:
                cell = ws.cell(row=r, column=col)
                estado = None
                if cell.value:
                    estado = str(cell.value).strip()
                elif cell.fill and cell.fill.fill_type == "solid":
                    estado = argb_to_hex(cell.fill.fgColor.rgb)
                if estado:
                    registros.append({
                        "empresa": empresa,
                        "pais": current_country,
                        "impuesto": current_tax,
                        "fecha": target_date or date_map[col],
                        "estado": estado,
                    })

    wb.close()
    return registros
