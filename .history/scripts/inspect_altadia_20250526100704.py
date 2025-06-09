#!/usr/bin/env python3
import sys, os
from openpyxl import load_workbook

# Configura tu ruta
EXCEL = "tax_calendar_25.xlsm"
wb = load_workbook(EXCEL, data_only=True)
ws = wb["ALTADIA"]

# Imprimimos toda la fila 1 y fila 5 (hasta la columna 50) para ver dónde está May-2025 y los días
print("Fila 1 (meses), cols 1-50:")
print([ws.cell(row=1, column=c).value for c in range(1, 51)])

print("\nFila 2 (años), cols 1-50:")
print([ws.cell(row=2, column=c).value for c in range(1, 51)])

print("\nFila 5 (días), cols 1-50:")
print([ws.cell(row=5, column=c).value for c in range(1, 51)])

wb.close()
