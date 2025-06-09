#!/usr/bin/env python3
import os
from openpyxl import load_workbook

EXCEL = os.path.join(os.path.dirname(os.path.dirname(__file__)), "tax_calendar_25.xlsm")
wb = load_workbook(EXCEL, data_only=True)
ws = wb["ALTADIA"]

max_col = ws.max_column
print(f"Total columnas detectadas: {max_col}\n")

print("=== Month headers (fila 1) ===")
for col in range(1, max_col + 1):
    val = ws.cell(row=1, column=col).value
    if isinstance(val, str) and " - " in val:
        print(f"Col {col}: {val}")

print("\n=== Days (fila 5) ===")
for col in range(1, max_col + 1):
    val = ws.cell(row=5, column=col).value
    if isinstance(val, (int, float)):
        print(f"Col {col}: {int(val)}")

wb.close()
