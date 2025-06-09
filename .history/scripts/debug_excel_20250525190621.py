"""
Módulo de lectura de Excel: parse_calendar lee las hojas de empresa y extrae los registros diarios de impuestos.
"""
from openpyxl import load_workbook
from typing import List, Dict, Any

EMPRESA_OVERRIDES = {
    "France": "Repsol",
    "Netherlands": "Repsol",
}


def argb_to_hex(argb: Any) -> str:
    if not argb:
        return None
    return f"#{str(argb)[-6:]}"


def parse_calendar(path_excel: str) -> List[Dict[str, Any]]:
    """
    Lee el fichero Excel y devuelve una lista de registros:
      empresa, pais, impuesto, fecha (date), estado (sigla o color hex).
    Ignora hojas 'SETTINGS' y las que empiezan 'CALENDAR'.
    """
    from datetime import datetime

    wb = load_workbook(path_excel, data_only=True)
    registros: List[Dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        # Saltar configuraciones o calendarios generales
        if sheet_name == 'SETTINGS' or sheet_name.startswith('CALENDAR'):
            continue
        ws = wb[sheet_name]
        empresa = EMPRESA_OVERRIDES.get(sheet_name, sheet_name)

        # 1) Buscar fila de 'Month - Year' en primeras 10 filas
        month_row = None
        month_info: Dict[int, tuple] = {}
        for r in range(1, 11):
            for cell in ws[r]:
                if isinstance(cell.value, str) and ' - ' in cell.value:
                    month_row = r
            if month_row:
                break
        if not month_row:
            continue
        # Construir misc de month per columna
        current_my = None
        for cell in ws[month_row]:
            if cell.value:
                parts = str(cell.value).split(' - ')
                if len(parts) == 2:
                    current_my = (parts[0], int(parts[1]))
            if current_my:
                month_info[cell.column] = current_my

        # 2) Buscar fila de días (valores 1-31)
        day_row = None
        for r in range(month_row + 1, month_row + 6):
            vals = [c.value for c in ws[r] if isinstance(c.value, (int, float))]
            if any(1 <= v <= 31 for v in vals):
                day_row = r
                break
        if not day_row:
            continue
        # Mapear columnas a fechas
        date_map: Dict[int, Any] = {}
        for cell in ws[day_row]:
            if isinstance(cell.value, (int, float)) and 1 <= int(cell.value) <= 31:
                col = cell.column
                day = int(cell.value)
                if col in month_info:
                    month_name, year = month_info[col]
                    fecha = datetime.strptime(f"{day} {month_name} {year}", "%d %B %Y").date()
                    date_map[col] = fecha

        # 3) Recorrer filas de datos desde day_row+1
        current_country = None
        current_tax = None
        for row in ws.iter_rows(min_row=day_row + 1):
            cellA = row[0]
            cellB = row[1]
            # País en col A
            if cellA.value and isinstance(cellA.value, str):
                current_country = cellA.value.strip()
            # Impuesto en col B
            if cellB.value and isinstance(cellB.value, str):
                current_tax = cellB.value.strip()
            if not current_country or not current_tax:
                continue
            # Celdas de fechas a partir de col C
            for cell in row[2:]:
                col = cell.column
                if col not in date_map:
                    continue
                # Determinar estado
                estado = None
                if cell.value:
                    estado = str(cell.value).strip()
                elif cell.fill and cell.fill.fill_type == 'solid':
                    estado = argb_to_hex(getattr(cell.fill.fgColor, 'rgb', None))
                if estado:
                    registros.append({
                        'empresa': empresa,
                        'pais': current_country,
                        'impuesto': current_tax,
                        'fecha': date_map[col],
                        'estado': estado,
                    })
    wb.close()
    return registros
