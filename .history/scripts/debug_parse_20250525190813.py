#!/usr/bin/env python3
from src.reader import parse_calendar
from openpyxl import load_workbook
from datetime import datetime

EXCEL_PATH = "tax_calendar_25.xlsm"

def debug_sheet(sheet_name):
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[sheet_name]
    print(f"\n=== Hoja: {sheet_name} ===")

    # 1) Encontrar fila de Mes-Año
    month_row = None
    for r in range(1, 11):
        for cell in ws[r]:
            if isinstance(cell.value, str) and " - " in cell.value:
                month_row = r
                mesanio = cell.value
                break
        if month_row:
            break
    print("  month_row:", month_row, "(", mesanio if month_row else "", ")")

    # 2) Mostrar mesinfo para col C-J
    if month_row:
        print("  month_info cols C-J:")
        for col in range(3, 11):
            val = ws.cell(row=month_row, column=col).value
            print(f"    Col {col}: {val}")

    # 3) Detectar day_row
    day_row = None
    for r in range((month_row or 1) + 1, (month_row or 1) + 8):
        vals = [c.value for c in ws[r] if isinstance(c.value, (int, float))]
        if any(1 <= v <= 31 for v in vals):
            day_row = r
            break
    print("  day_row:", day_row)
    if day_row:
        print("  días cols C-J:", [ws.cell(row=day_row, column=c).value for c in range(3, 11)])

    wb.close()

def main():
    wb = load_workbook(EXCEL_PATH, data_only=True)
    # Haz debug solo de unas pocas hojas representativas:
    for sheet in ["ENDESA", "DRAGADOS", "X-ELIO"]:
        if sheet in wb.sheetnames:
            debug_sheet(sheet)
    wb.close()

    # Por último, invoca parse_calendar y ve cuántos registros arroja
    regs = parse_calendar(EXCEL_PATH)
    print("\nTotal registros parseados:", len(regs))
    print("Primeros 10 (si hay):")
    for r in regs[:10]:
        print(f"  {r}")

if __name__ == "__main__":
    main()
