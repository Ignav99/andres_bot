from openpyxl import load_workbook
from datetime import datetime, date
from typing import List, Dict, Any, Optional
import re

def argb_to_hex(argb: Any) -> Optional[str]:
    """
    Convierte un valor ARGB (o similar) a '#RRGGBB'.
    """
    if not argb:
        return None
    s = str(argb)
    return f"#{s[-6:]}"

# Mapear hojas especiales a la empresa correcta
EMPRESA_OVERRIDES = {
    "France": "Repsol",
    "Netherlands": "Repsol",
}

def parse_calendar(
    path_excel: str,
    target_date: Optional[date] = None
) -> List[Dict[str, Any]]:
    """
    Lee el fichero Excel y devuelve una lista de registros con:
      empresa, pais, impuesto, fecha, estado (texto o color hex).

    Si target_date se pasa, solo extrae registros para esa fecha concreta.
    Ignora las hojas 'SETTINGS' y las que empiezan por 'CALENDAR'.
    """
    wb = load_workbook(path_excel, data_only=True)
    registros: List[Dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "SETTINGS" or sheet_name.startswith("CALENDAR"):
            continue

        ws = wb[sheet_name]
        empresa = EMPRESA_OVERRIDES.get(sheet_name, sheet_name)

        # 1) Encontrar fila de Mes-Año (buscamos "Month - YYYY")
        month_row = None
        for r in range(1, 11):
            for cell in ws[r]:
                val = cell.value
                if isinstance(val, str) and re.match(r"^[A-Za-z]+ - \d{4}$", val.strip()):
                    month_row = r
                    break
            if month_row:
                break
        if not month_row:
            continue  # sin cabecera válida

        # 2) Construir map columna→(mes, año), propagando el último válido
        month_info: Dict[int, tuple] = {}
        current_my = None
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=month_row, column=col).value
            if isinstance(val, str) and re.match(r"^[A-Za-z]+ - \d{4}$", val.strip()):
                mes_str, anyo_str = val.split(" - ")
                try:
                    anyo = int(anyo_str)
                    current_my = (mes_str, anyo)
                except ValueError:
                    current_my = None
            if current_my:
                month_info[col] = current_my

        # 3) Detectar fila de días: forzamos la fila 5 donde están 1,2,3...
        day_row = 5
        # validación
        nums = [c.value for c in ws[day_row] if isinstance(c.value, (int, float))]
        if not (len(set(nums)) >= 2 and all(1 <= int(v) <= 31 for v in nums)):
            continue

        # 4) Si hay target_date, buscamos su columna y saltamos si no existe
        col_for_date = None
        if target_date:
            for col, (mes, anyo) in month_info.items():
                cell = ws.cell(row=day_row, column=col)
                if isinstance(cell.value, (int, float)):
                    dia = int(cell.value)
                    try:
                        fecha = datetime.strptime(f"{dia} {mes} {anyo}", "%d %B %Y").date()
                    except ValueError:
                        continue
                    if fecha == target_date:
                        col_for_date = col
                        break
            if col_for_date is None:
                continue  # esta hoja no tiene ese día

        # 5) Mapear columnas a fechas (solo si no hay target_date)
        date_map: Dict[int, date] = {}
        if not target_date:
            for col, (mes, anyo) in month_info.items():
                cell = ws.cell(row=day_row, column=col)
                if isinstance(cell.value, (int, float)):
                    dia = int(cell.value)
                    try:
                        date_map[col] = datetime.strptime(f"{dia} {mes} {anyo}", "%d %B %Y").date()
                    except ValueError:
                        pass

        # 6) Recorrer filas de datos
        current_country = None
        current_tax = None
        for row in ws.iter_rows(min_row=day_row + 1):
            cellA = row[0]  # columna A → país
            if isinstance(cellA.value, str) and cellA.value.strip().lower().startswith("legend"):
                break  # no procesar leyenda

            cellB = row[1]  # columna B → impuesto
            if isinstance(cellA.value, str) and cellA.value.strip():
                current_country = cellA.value.strip()
            if isinstance(cellB.value, str) and cellB.value.strip():
                current_tax = cellB.value.strip()
            if not current_country or not current_tax:
                continue

            # 7) Extraer estado solo de la columna objetivo (o todas si no hay target_date)
            if target_date:
                cols = [col_for_date]
            else:
                cols = list(date_map.keys())

            for col in cols:
                # row es un tuple, 0-based, col es 1-based
                cell = row[col - 1]
                estado = None
                if cell.value:
                    estado = str(cell.value).strip()
                elif cell.fill and cell.fill.fill_type == "solid":
                    estado = argb_to_hex(getattr(cell.fill.fgColor, "rgb", None))
                if estado:
                    registros.append({
                        "empresa": empresa,
                        "pais": current_country,
                        "impuesto": current_tax,
                        "fecha": target_date or date_map[col],
                        "estado": estado,
                    })

    wb.close()
    return registros
