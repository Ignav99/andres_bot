from openpyxl import load_workbook
from datetime import datetime, date
from typing import List, Dict, Any, Optional
import re

def argb_to_hex(argb) -> Optional[str]:
    if not argb:
        return None
    s = str(argb)
    return f"#{s[-6:]}"

# Mapa de color de fondo (hex) a código
COLOR_CODE = {
    "#FFFF66": "SI",
    "#9966FF": "RI",
    "#FF5050": "SD",
    "#FF99FF": "AD",
    "#70AD47": "OS",
    "#00B0F0": "OP",
    "#996633": "HS",
    "#BF8F00": "HL",
}

# Leyenda de códigos
LEGEND = {
    "SI": "Send information",
    "RI": "Review information and doubts (EY Local)",
    "SD": "Send draft (EY Local)",
    "AD": "Approve draft",
    "OS": "Official Submission Deadline",
    "OP": "Official Payment Deadline",
    "SP": "Official Submission and Payment (same deadline)",
    "HS": "Public Holiday Spain",
    "HL": "Local Holiday – Non-working day"
}

EMPRESA_OVERRIDES = {
    "France": "Repsol",
    "Netherlands": "Repsol",
}

def parse_calendar(path_excel: str, target_date: Optional[date]=None) -> List[Dict[str, Any]]:
    wb = load_workbook(path_excel, data_only=True)
    out: List[Dict[str, Any]] = []

    for sheet in wb.sheetnames:
        if sheet == "SETTINGS" or sheet.startswith("CALENDAR"):
            continue
        ws = wb[sheet]
        empresa = EMPRESA_OVERRIDES.get(sheet, sheet)

        # 1) Detectar fila de mes-año
        month_row = next((r for r in range(1,6)
                          if any(isinstance(c.value,str) and re.match(r"^[A-Za-z]+ - \d{4}$", c.value) for c in ws[r])),
                         None)
        if not month_row:
            continue

        # 2) Mapear mes→columna base
        month_headers: Dict[tuple,int] = {}
        for col in range(1, ws.max_column+1):
            val = ws.cell(row=month_row, column=col).value
            if isinstance(val, str):
                m = re.match(r"^([A-Za-z]+) - (\d{4})$", val)
                if m:
                    month_headers[(m.group(1).lower(), int(m.group(2)))] = col

        # 3) Columnas objetivo y mapa col→fecha
        cols: List[int] = []
        date_map: Dict[int,date] = {}

        if target_date:
            key = (target_date.strftime("%B").lower(), target_date.year)
            base = month_headers.get(key)
            if not base:
                continue
            col_t = base + (target_date.day - 1)
            cols = [col_t]
            date_map[col_t] = target_date
        else:
            from datetime import timedelta
            for (mes, anyo), base in month_headers.items():
                try:
                    month_num = datetime.strptime(mes, "%B").month
                except:
                    continue
                for d in range(1,32):
                    try:
                        dt = date(anyo, month_num, d)
                    except ValueError:
                        break
                    col = base + (d-1)
                    if col > ws.max_column:
                        break
                    cols.append(col)
                    date_map[col] = dt

        # 4) Recorrer filas de datos desde 6 hasta la leyenda
        for r in range(6, ws.max_row+1):
            a = ws.cell(row=r, column=1).value
            if isinstance(a, str) and a.strip().lower().startswith("legend"):
                break
            b = ws.cell(row=r, column=2).value
            if not (isinstance(a,str) and a.strip() and isinstance(b,str) and b.strip()):
                continue
            pais = a.strip()
            imp  = b.strip()

                        # 5) Solo celdas de interés
            for col in cols:
                    cell = ws.cell(row=r, column=col)
                    estado = None

                    # 1) texto directo
                    val = cell.value
                    if isinstance(val, str) and val.strip().upper() in LEGEND:
                        estado = val.strip().upper()
                    else:
                        # 2) color de fondo sólido
                        fill = cell.fill
                        if fill and fill.fill_type == "solid":
                            # primero probamos fgColor.rgb
                            raw = getattr(fill.fgColor, "rgb", None)
                            if not raw:
                                # si no hay rgb, probamos indexed
                                raw = getattr(fill.fgColor, "indexed", None)
                            hexc = argb_to_hex(raw)
                            # debug: muestro el hex y la fila/col
                            print(f"[DEBUG color] hoja={sheet} fila={r} col={col} hex={hexc}")
                            estado = COLOR_CODE.get(hexc)

                    if estado:
                        out.append({
                            "empresa": empresa,
                            "pais": pais,
                            "impuesto": imp,
                            "fecha": date_map[col],
                            "estado": estado
                        })


    wb.close()
    return out
