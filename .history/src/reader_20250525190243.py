from openpyxl import load_workbook
from datetime import datetime
from typing import List, Dict, Any

def argb_to_hex(argb: str) -> str:
    return f"#{argb[-6:]}" if argb else None

EMPRESA_OVERRIDES = {
    "France": "Repsol",
    "Netherlands": "Repsol"
}

def parse_calendar(path_excel: str) -> List[Dict[str, Any]]:
    wb = load_workbook(path_excel, data_only=True)
    registros = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        empresa = EMPRESA_OVERRIDES.get(sheet_name, sheet_name)

        # 1) Construir map de columna→fecha desde fila 2 (mes-año) y fila 3 (día)
        fecha_map = {}
        for cell in ws[3]:
            try:
                dia = int(cell.value)
                mes_anio = ws.cell(row=2, column=cell.column).value  # e.g. "June-2025"
                fecha = datetime.strptime(f"{dia} {mes_anio}", "%d %B-%Y")
                fecha_map[cell.column] = fecha
            except Exception:
                continue

        current_country = None
        # 2) Recorremos filas a partir de la 4
        for row in ws.iter_rows(min_row=4):
            cell_B = row[1]  # índice 1 == columna B

            # 2.1) País: celda gris (código aproximado FFCCCCCC)
            # Obtener el RGB como cadena (si existe)
            rgb_obj = getattr(cell_B.fill.fgColor, 'rgb', None)
            rgb_str = str(rgb_obj) if rgb_obj is not None else ''
            # Detección de gris claro (código empieza por "FFCC")
            if cell_B.fill and cell_B.fill.fill_type == 'solid' and rgb_str.upper().startswith("FFCC"):
                current_country = cell_B.value
                continue


            # 2.2) Si B tiene texto, es un tipo de impuesto
            if cell_B.value:
                current_tax = cell_B.value

                # 2.3) Recorremos celdas de calendario
                for cell in row:
                    if cell.column not in fecha_map:
                        continue
                    estado = None
                    if cell.value:
                        estado = str(cell.value).strip()
                    elif (cell.fill and cell.fill.fill_type=='solid'):
                        estado = argb_to_hex(cell.fill.fgColor.rgb)
                    if estado:
                        registros.append({
                            "empresa": empresa,
                            "pais": current_country,
                            "impuesto": current_tax,
                            "fecha": fecha_map[cell.column],
                            "estado": estado
                        })
    wb.close()
    return registros
