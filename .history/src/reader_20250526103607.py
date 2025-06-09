from openpyxl import load_workbook
from datetime import datetime, date
from typing import List, Dict, Any, Optional
import re

def argb_to_hex(argb) -> Optional[str]:
    if not argb:
        return None
    s = str(argb)
    return f"#{s[-6:]}"

EMPRESA_OVERRIDES = {
    "France": "Repsol",
    "Netherlands": "Repsol",
}

def parse_calendar(
    path_excel: str,
    target_date: Optional[date] = None
) -> List[Dict[str, Any]]:
    wb = load_workbook(path_excel, data_only=True)
    out: List[Dict[str, Any]] = []

    for sheet in wb.sheetnames:
        if sheet in ("SETTINGS",) or sheet.startswith("CALENDAR"):
            continue
        ws = wb[sheet]
        empresa = EMPRESA_OVERRIDES.get(sheet, sheet)

        # 1) Detectar mes_row (fila con "Month - YYYY")
        month_row = None
        for r in range(1, 6):
            for cell in ws[r]:
                if isinstance(cell.value, str) and re.match(r"^[A-Za-z]+ - \d{4}$", cell.value.strip()):
                    month_row = r
                    break
            if month_row:
                break
        if not month_row:
            continue

        # 2) Mapear encabezados de mes → columna inicial
        month_headers: Dict[tuple, int] = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=month_row, column=col).value
            if isinstance(val, str):
                m = re.match(r"^([A-Za-z]+) - (\d{4})$", val.strip())
                if m:
                    mes_str, anyo_str = m.group(1), m.group(2)
                    month_headers[(mes_str.lower(), int(anyo_str))] = col

        # 3) Determinar columnas a chequear
        if target_date:
            # Buscar encabezado para el mes/año de target_date
            key = (target_date.strftime("%B").lower(), target_date.year)
            base_col = month_headers.get(key)
            if base_col is None:
                continue
            # columna exacta = base + (día - 1)
            col_target = base_col + (target_date.day - 1)
            if not (3 <= col_target <= ws.max_column):
                continue
            cols_to_check = [col_target]
            date_map = {col_target: target_date}
        else:
            # Todas las fechas: construimos date_map de mes en mes
            from datetime import timedelta
            cols_to_check = []
            date_map = {}
            for (mes, anyo), base_col in month_headers.items():
                try:
                    month_num = datetime.strptime(mes, "%B").month
                except:
                    continue
                for day in range(1, 32):
                    try:
                        dt = date(anyo, month_num, day)
                    except ValueError:
                        break
                    col = base_col + (day - 1)
                    if col > ws.max_column:
                        break
                    cols_to_check.append(col)
                    date_map[col] = dt

        # 4) Recorrer filas de datos (a partir de row 6) hasta "Legend"
        for r in range(6, ws.max_row + 1):
            pais = ws.cell(row=r, column=1).value
            if isinstance(pais, str) and pais.strip().lower().startswith("legend"):
                break
            impuesto = ws.cell(row=r, column=2).value
            if not (isinstance(pais, str) and pais.strip() and isinstance(impuesto, str) and impuesto.strip()):
                continue
            pais = pais.strip()
            impuesto = impuesto.strip()

            # 5) Solo la(s) columna(s) de interés
            for col in cols_to_check:
                cell = ws.cell(row=r, column=col)
                estado = None
                if cell.value:
                    estado = str(cell.value).strip()
                elif cell.fill and cell.fill.fill_type == "solid":
                    estado = argb_to_hex(cell.fill.fgColor.rgb)
                if estado:
                    out.append({
                        "empresa": empresa,
                        "pais": pais,
                        "impuesto": impuesto,
                        "fecha": date_map[col],
                        "estado": estado,
                    })
    wb.close()
    return out
