from openpyxl import load_workbook
from datetime import datetime, date
from typing import List, Dict, Any, Optional
import re

def argb_to_hex(argb) -> Optional[str]:
    if not argb:
        return None
    s = str(argb)
    return f"#{s[-6:]}"

# Mapa de color de fondo (hex) a código
COLOR_CODE = {
    "#FFFF66": "SI",
    "#9966FF": "RI",
    "#FF5050": "SD",
    "#FF99FF": "AD",
    "#70AD47": "OS",
    "#00B0F0": "OP",
    "#996633": "HS",
    "#BF8F00": "HL",
    "#CC9900": "HL",
}

# Leyenda de códigos
LEGEND = {
    "SI": "Send information",
    "RI": "Review information and doubts (EY Local)",
    "SD": "Send draft (EY Local)",
    "AD": "Approve draft",
    "OS": "Official Submission Deadline",
    "OP": "Official Payment Deadline",
    "SP": "Official Submission and Payment (same deadline)",
    "HS": "Public Holiday Spain",
    "HL": "Local Holiday – Non-working day"
}

EMPRESA_OVERRIDES = {
    "France": "Repsol",
    "Netherlands": "Repsol",
}

def parse_calendar(
    path_excel: str,
    target_date: Optional[date] = None,
    company_filter: Optional[str] = None
) -> List[Dict[str, Any]]:
    wb = load_workbook(path_excel, data_only=True)
    registros: List[Dict[str, Any]] = []

    for sheet in wb.sheetnames:
        if sheet == "SETTINGS" or sheet.startswith("CALENDAR"):
            continue

        empresa = EMPRESA_OVERRIDES.get(sheet, sheet)
        # filtrado por empresa
        if company_filter and empresa.lower() != company_filter.lower():
            continue

        ws = wb[sheet]
        print(f"\n[DEBUG] Hoja: {sheet}, Empresa: {empresa}")

        # 1) Detectar fila de mes-año: filas 1–5, o 2–6 si es Repsol
        start_row = 2 if empresa.lower() == "repsol" else 1
        month_row = None
        for r in range(start_row, start_row + 5):
            for c in ws[r]:
                val = c.value
                if isinstance(val, str) and "-" in val and re.search(r"\b\d{4}\b", val):
                    month_row = r
                    break
            if month_row:
                break
        print(f"[DEBUG]   month_row = {month_row}")
        if not month_row:
            continue

        # 2) Mapear mes→columna base
        month_headers: Dict[tuple,int] = {}
        for col in range(1, ws.max_column+1):
            val = ws.cell(row=month_row, column=col).value
            if isinstance(val, str):
                m = re.match(r"^([A-Za-z]+)\s*-\s*(\d{4})$", val.strip())
                if m:
                    mes = m.group(1).lower()
                    anio = int(m.group(2))
                    month_headers[(mes, anio)] = col
        print(f"[DEBUG]   month_headers keys: {list(month_headers.keys())}")

        # 3) Determinar columna objetivo
        cols: List[int] = []
        date_map: Dict[int,date] = {}
        if target_date:
            key = (target_date.strftime("%B").lower(), target_date.year)
            base = month_headers.get(key)
            print(f"[DEBUG]   buscando key {key} -> base = {base}")
            if not base:
                continue
            col_t = base + (target_date.day - 1)
            print(f"[DEBUG]   target_date {target_date} -> col_target = {col_t}")
            cols = [col_t]
            date_map[col_t] = target_date
        else:
            # (rama --all-dates si la necesitas)
            continue

        # 4) Recorrer filas de datos
        for r in range(6, ws.max_row+1):
            pa = ws.cell(row=r, column=1).value
            if isinstance(pa, str) and pa.strip().lower().startswith("legend"):
                break
            ip = ws.cell(row=r, column=2).value
            if not (isinstance(pa,str) and pa.strip() and isinstance(ip,str) and ip.strip()):
                continue
            pais = pa.strip()
            imp  = ip.strip()

            for col in cols:
                cell = ws.cell(row=r, column=col)
                estado = None

                # A) Por texto (siglas)
                val = cell.value
                if isinstance(val, str) and val.strip().upper() in LEGEND:
                    estado = val.strip().upper()
                    print(f"[DEBUG]     reconocido texto -> {estado}")
                else:
                    # B) Por color de fondo
                    fill = cell.fill
                    raw = None
                    if fill and fill.fill_type == "solid":
                        raw = getattr(fill.start_color, "rgb", None) or getattr(fill.fgColor, "rgb", None)
                    if raw:
                        hexc = argb_to_hex(raw)
                        print(f"[DEBUG]     color hex detectado -> {hexc}")
                        estado = COLOR_CODE.get(hexc)
                        print(f"[DEBUG]     mapeado a estado -> {estado}")

                if estado:
                    registros.append({
                        "empresa": empresa,
                        "pais": pais,
                        "impuesto": imp,
                        "fecha": date_map[col],
                        "estado": estado
                    })

    wb.close()
    return registros

