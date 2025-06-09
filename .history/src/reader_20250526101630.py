from openpyxl import load_workbook
from datetime import datetime, date, timedelta
from typing import List, Dict, Any, Optional

def argb_to_hex(argb) -> Optional[str]:
    if not argb: return None
    s = str(argb); return f"#{s[-6:]}"

EMPRESA_OVERRIDES = {"France":"Repsol","Netherlands":"Repsol"}

def parse_calendar(path_excel: str, target_date: Optional[date]=None) -> List[Dict[str, Any]]:
    wb = load_workbook(path_excel, data_only=True)
    out: List[Dict[str, Any]] = []

    for sheet in wb.sheetnames:
        if sheet=="SETTINGS" or sheet.startswith("CALENDAR"): continue
        ws = wb[sheet]
        emp = EMPRESA_OVERRIDES.get(sheet, sheet)

        # Año (fila 2, columna C)
        year_cell = ws.cell(row=2, column=3).value
        try:
            year = int(year_cell)
        except:
            continue

        # Determinar lista de columnas a chequear
        if target_date:
            if target_date.year != year:
                continue
            # columna = 3 + (fecha - 1 ene).days
            delta = (target_date - date(year,1,1)).days
            col_target = 3 + delta
            if col_target < 3 or col_target > ws.max_column:
                continue
            cols_to_check = [col_target]
            fechas_map = {col_target: target_date}
        else:
            # si quisieras todo el año, puedes crear date_map así:
            cols_to_check = []
            fechas_map = {}
            for col in range(3, ws.max_column+1):
                d = date(year,1,1) + timedelta(days=col-3)
                cols_to_check.append(col)
                fechas_map[col] = d

        # fila donde empiezan datos (empieza en A/B) y antes de leyenda
        # asumimos datos desde row=6 (tras cabeceras y días)
        for r in range(6, ws.max_row+1):
            a = ws.cell(row=r, column=1).value
            if isinstance(a,str) and a.strip().lower().startswith("legend"):
                break
            b = ws.cell(row=r, column=2).value
            if not (isinstance(a,str) and a.strip() and isinstance(b,str) and b.strip()):
                continue
            pais = a.strip(); imp = b.strip()

            for col in cols_to_check:
                cell = ws.cell(row=r, column=col)
                estado = None
                if cell.value:
                    estado = str(cell.value).strip()
                elif cell.fill and cell.fill.fill_type=="solid":
                    estado = argb_to_hex(cell.fill.fgColor.rgb)
                if estado:
                    out.append({
                        "empresa": emp,
                        "pais": pais,
                        "impuesto": imp,
                        "fecha": fechas_map[col],
                        "estado": estado
                    })
    wb.close()
    return out
