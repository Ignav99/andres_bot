from openpyxl import load_workbook
from datetime import datetime, date
from typing import List, Dict, Any, Optional

LEGEND_ROWS_START_TEXT = ("legend",)  # para detectar inicio de leyenda

def argb_to_hex(argb) -> Optional[str]:
    if not argb: return None
    s = str(argb)
    return f"#{s[-6:]}"

EMPRESA_OVERRIDES = {"France":"Repsol", "Netherlands":"Repsol"}

def parse_calendar(path: str, target_date: Optional[date]=None) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    out = []

    for sheet in wb.sheetnames:
        if sheet=="SETTINGS" or sheet.startswith("CALENDAR"): continue
        ws = wb[sheet]
        emp = EMPRESA_OVERRIDES.get(sheet, sheet)

        # 1) Encontrar la columna EXACTA para target_date
        col_target = None
        if target_date:
            # Recorremos las (digamos) primeras 5 filas que contienen cabeceras
            for header_row in range(1,6):
                for cell in ws[header_row]:
                    val = cell.value
                    # Puede estar en formato texto "6" y la fila anterior dice "May-2025",
                    # o puede estar todo unido "06/05/2025". Probemos ambas:
                    # A) Si la celda es fecha: 
                    if isinstance(val, datetime):
                        if val.date()==target_date:
                            col_target = cell.column
                            break
                    # B) Si es número y la fila de arriba (header_row-1) tiene month-year:
                    if isinstance(val, (int,float)):
                        # leer mes-año justo encima
                        mh = ws.cell(row=header_row-1, column=cell.column).value
                        if isinstance(mh, str) and target_date.strftime("%B").lower() in mh.lower() \
                           and str(target_date.year) in mh \
                           and int(val)==target_date.day:
                            col_target = cell.column
                            break
                if col_target: break
            if not col_target:
                continue  # esta hoja no tiene esa fecha

        # 2) Recorrer filas **solo** en esa columna
        #    desde la fila donde empieza el país/impuesto (detectarla),
        #    hasta la fila antes de que aparezca la leyenda.
        # Buscamos la primera fila que contenga un país real
        start_row = None
        for r in range(1, ws.max_row+1):
            a = ws.cell(row=r, column=1).value
            if isinstance(a,str) and a.strip() and a.strip().lower() not in LEGEND_ROWS_START_TEXT:
                # Primera vez que vemos algo en col A que no es cabecera ni leyenda...
                # asegurémonos de que la fila  contiene también algo en B:
                if ws.cell(row=r, column=2).value:
                    start_row = r
                    break
        if not start_row:
            continue

        # Recorremos hasta encontrar la fila donde A sea 'legend' (o hasta max)
        for r in range(start_row, ws.max_row+1):
            a = ws.cell(row=r, column=1).value
            if isinstance(a,str) and a.strip().lower().startswith("legend"):
                break
            # en cada fila, si la celda de col_target tiene color o valor:
            c = ws.cell(row=r, column=col_target)
            estado = None
            if c.value:
                estado = str(c.value).strip()
            elif c.fill and c.fill.fill_type=='solid':
                estado = argb_to_hex(c.fill.fgColor.rgb)
            if estado:
                pais = ws.cell(row=r, column=1).value.strip()
                imp  = ws.cell(row=r, column=2).value.strip()
                out.append({
                    "empresa": emp,
                    "pais": pais,
                    "impuesto": imp,
                    "fecha": target_date,
                    "estado": estado
                })
    wb.close()
    return out
