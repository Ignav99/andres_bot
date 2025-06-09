from openpyxl import load_workbook
from datetime import datetime
from typing import List, Dict, Any

def argb_to_hex(argb: str) -> str:
    return f"#{argb[-6:]}" if argb else None

# Mapeo especial de pestañas a empresas
EMPRESA_OVERRIDES = {
    "France": "Repsol",
    "Netherlands": "Repsol"
}

def parse_calendar(path_excel: str) -> List[Dict[str, Any]]:
    wb = load_workbook(path_excel, data_only=True)
    registros = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 1) Empresa
        empresa = EMPRESA_OVERRIDES.get(sheet_name, sheet_name)

        # 2) Detectar fila de fechas (suponemos que es la fila 3)
        fecha_map = {}
        header_row = 3
        for cell in ws[header_row]:
            try:
                # Extraer día del mes (p.ej. '21') y mes+año de otro header superior si hiciese falta
                dia = int(cell.value)
                # Asumimos que la celda superior (fila 2) contiene el mes y año, tipo "June-2025"
                mes_anio = ws.cell(row=2, column=cell.column).value  # ej. "June-2025"
                fecha = datetime.strptime(f"{dia} {mes_anio}", "%d %B-%Y")
                fecha_map[cell.column] = fecha
            except Exception:
                continue

        current_country = None
        # 3) Iterar filas desde la 4 hacia abajo
        for row in ws.iter_rows(min_row=4):
            cell_B = row[1]  # columna B, índice 1
            # detectar país por color de relleno gris
            if cell_B.fill and cell_B.fill.fill_type == 'solid' and cell_B.fill.fgColor.rgb.startswith("FFCCCCCC"):
                current_country = cell_B.value
                continue

            # si no es cabecera de país y B tiene texto, es impuesto
            if cell_B.value:
                current_tax = cell_B.value

                # 4) recorrer columnas de fecha
                for cell in row:
                    col = cell.column
                    if col not in fecha_map:
                        continue
                    estado = None
                    # preferimos sigla en valor
                    if cell.value:
                        estado = str(cell.value).strip()
                    # si no, buscamos color
                    elif cell.fill and cell.fill.fill_type == 'solid':
                        estado = argb_to_hex(cell.fill.fgColor.rgb)
                    if estado:
                        registros.append({
                            "empresa": empresa,
                            "pais": current_country,
                            "impuesto": current_tax,
                            "fecha": fecha_map[col],
                            "estado": estado
                        })

    wb.close()
    return registros
