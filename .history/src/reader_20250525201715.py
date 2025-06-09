from openpyxl import load_workbook
from datetime import datetime
from typing import List, Dict, Any
import re

def argb_to_hex(argb: Any) -> str:
    """
    Convierte un valor ARGB (o similar) a '#RRGGBB'.
    """
    if not argb:
        return None
    s = str(argb)
    return f"#{s[-6:]}"

# Mapear hojas especiales a la empresa correcta
EMPRESA_OVERRIDES = {
    "France": "Repsol",
    "Netherlands": "Repsol",
}

def parse_calendar(path_excel: str) -> List[Dict[str, Any]]:
    """
    Lee el fichero Excel y devuelve una lista de registros con:
        empresa, pais, impuesto, fecha (date), estado (texto o color hex).
    Ignora las hojas 'SETTINGS' y las que empiezan por 'CALENDAR'.
    """
    wb = load_workbook(path_excel, data_only=True)
    registros: List[Dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "SETTINGS" or sheet_name.startswith("CALENDAR"):
            continue

        ws = wb[sheet_name]
        empresa = EMPRESA_OVERRIDES.get(sheet_name, sheet_name)

        # 1) Encontrar fila de Mes-Año (buscamos "Month - YYYY")
        month_row = None
        for r in range(1, 11):
            for cell in ws[r]:
                val = cell.value
                if isinstance(val, str) and re.match(r"^[A-Za-z]+ - \d{4}$", val.strip()):
                    month_row = r
                    break
            if month_row:
                break
        if not month_row:
            # No encontramos cabecera de mes-año → saltamos hoja
            continue

        # 2) Construir map columna→(mes, año), propagando el último válido
        month_info: Dict[int, tuple] = {}
        current_my = None
        max_col = ws.max_column
        for col in range(1, max_col + 1):
            val = ws.cell(row=month_row, column=col).value
            if isinstance(val, str) and re.match(r"^[A-Za-z]+ - \d{4}$", val.strip()):
                mes_str, anyo_str = val.split(" - ")
                try:
                    anyo = int(anyo_str)
                    current_my = (mes_str, anyo)
                except ValueError:
                    # Si no es un año válido, lo ignoramos
                    pass
            if current_my:
                month_info[col] = current_my

                # 3) Fijar la fila de días en la fila 5 (donde vienen 1,2,3,4...)
        day_row = 5
        print(f"[DEBUG reader] Hoja={sheet_name}, usando day_row={day_row}")  # <--- log
        # Validar: que esa fila contenga varios días distintos 1–31
        vals = [c.value for c in ws[day_row] if isinstance(c.value, (int, float))]
        print(f"[DEBUG reader] Hoja={sheet_name}, valores en fila 5: {vals}")  # <--- log
        if not (len(set(vals)) >= 3 and all(1 <= int(v) <= 31 for v in vals)):
            print(f"[DEBUG reader] Hoja={sheet_name}: fila 5 inválida, salto hoja")  # <--- log
            continue




        # 4) Mapear columnas a fecha concreta
        date_map: Dict[int, Any] = {}
        for col, (mes, anyo) in month_info.items():
            cell = ws.cell(row=day_row, column=col)
            if isinstance(cell.value, (int, float)):
                dia = int(cell.value)
                try:
                    fecha = datetime.strptime(f"{dia} {mes} {anyo}", "%d %B %Y").date()
                    date_map[col] = fecha
                except ValueError:
                    # Mes con nombre distinto al inglés → ajustar si hiciera falta
                    pass

        # 5) Recorrer filas de datos
        current_country = None
        current_tax = None
        for row in ws.iter_rows(min_row=day_row + 1):
            cellA = row[0]  # columna A → país
            # Si llegamos a la fila de leyenda (texto 'Legend' o similar), rompemos
            if isinstance(cellA.value, str) and cellA.value.strip().lower().startswith("legend"):
                break
            cellB = row[1]  # columna B → impuesto

            # País (col A)
            if isinstance(cellA.value, str) and cellA.value.strip():
                current_country = cellA.value.strip()

            # Impuesto (col B)
            if isinstance(cellB.value, str) and cellB.value.strip():
                current_tax = cellB.value.strip()

            if not current_country or not current_tax:
                continue

            # Para cada celda en columnas de fecha (>= 3):
            for cell in row[2:]:
                col = cell.column
                if col not in date_map:
                    continue

                estado = None
                if cell.value:
                    estado = str(cell.value).strip()
                elif cell.fill and cell.fill.fill_type == "solid":
                    estado = argb_to_hex(getattr(cell.fill.fgColor, "rgb", None))

                if estado:
                    registros.append({
                        "empresa": empresa,
                        "pais": current_country,
                        "impuesto": current_tax,
                        "fecha": date_map[col],
                        "estado": estado,
                    })

    wb.close()
    return registros
