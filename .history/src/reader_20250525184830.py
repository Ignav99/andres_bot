"""
Módulo de lectura de Excel: función leer_excel(path) que extrae celdas coloreadas con su metadata.
"""
from openpyxl import load_workbook
from typing import List, Dict

def argb_to_hex(argb: str) -> str:
    if not argb:
        return None
    return '#' + argb[-6:]


def leer_excel(path_excel: str) -> List[Dict]:
    """
    Lee un fichero Excel y devuelve una lista de registros con:
    [{'hoja': str, 'celda': str, 'valor': Any, 'color': str}, ...]
    Sólo incluye celdas con relleno sólido.
    """
    wb = load_workbook(filename=path_excel, data_only=True)
    registros = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                fill = cell.fill
                if fill and fill.fill_type == 'solid':
                    color_hex = argb_to_hex(fill.fgColor.rgb)
                    registros.append({
                        'hoja': sheet_name,
                        'celda': cell.coordinate,
                        'valor': cell.value,
                        'color': color_hex,
                    })
    wb.close()
    return registros
