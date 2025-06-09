#!/usr/bin/env python3
import sys, os

# Permitir import de src/
PROYECTO_ROOT = os.path.dirname(os.path.dirname(__file__))
sys.path.insert(0, PROYECTO_ROOT)

from openpyxl import load_workbook

EXCEL = os.path.join(PROYECTO_ROOT, "tax_calendar_25.xlsm")

def main():
    wb = load_workbook(EXCEL, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"\n--- Hoja: {sheet} ---")
        # Filas 1–10, columnas A–J
        for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10):
            print([cell.value for cell in row])
    wb.close()

if __name__ == "__main__":
    main()
