#!/usr/bin/env python3
import sys, os

# Importar parse_calendar
PROYECTO_ROOT = os.path.dirname(os.path.dirname(__file__))
sys.path.insert(0, PROYECTO_ROOT)

from src.reader import parse_calendar
from openpyxl import load_workbook
from datetime import datetime

EXCEL = os.path.join(PROYECTO_ROOT, "tax_calendar_25.xlsm")

def debug_sheet(name):
    wb = load_workbook(EXCEL, data_only=True)
    ws = wb[name]
    print(f"\n=== Hoja: {name} ===")

    # 1) month_row
    month_row = None; month_val = None
    for r in range(1, 11):
        for cell in ws[r]:
            if isinstance(cell.value, str) and " - " in cell.value:
                month_row = r; month_val = cell.value; break
        if month_row: break
    print(" month_row:", month_row, f"({month_val})")

    # 2) valores Mes-Año en cols 3–10
    if month_row:
        print(" month_info cols 3–10:", [ws.cell(row=month_row, column=c).value for c in range(3, 11)])

    # 3) day_row
    day_row = None
    for r in range((month_row or 1)+1, (month_row or 1)+8):
        nums = [c.value for c in ws[r] if isinstance(c.value, (int, float))]
        if len(set(nums))>1 and all(1<=int(v)<=31 for v in nums):
            day_row = r; break
    print(" day_row:", day_row)

    # 4) valores días cols 3–10
    if day_row:
        print(" days cols 3–10:", [ws.cell(row=day_row, column=c).value for c in range(3, 11)])

    wb.close()

def main():
    wb = load_workbook(EXCEL, data_only=True)
    for sheet in ["ENDESA", "DRAGADOS", "X-ELIO","ALTADIA", "REPSOL"]:
        if sheet in wb.sheetnames:
            debug_sheet(sheet)
    wb.close()

    regs = parse_calendar(EXCEL)
    print("\nTotal registros parseados:", len(regs))
    if regs:
        print("Primeros 10 registros:")
        for r in regs[:10]:
            print(" ", r)

if __name__ == "__main__":
    main()
